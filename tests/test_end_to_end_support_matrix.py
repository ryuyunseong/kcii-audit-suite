from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from kcii_audit.cli import app
from kcii_audit.reports.questionnaire_writer import QUESTIONNAIRE_SHEETS

runner = CliRunner()

SEVEN_OUTPUTS = {
    "evidence.jsonl",
    "results.json",
    "detail.xlsx",
    "summary.xlsx",
    "report.md",
    "security_advisory.md",
    "security_advisory.xlsx",
}

FIVE_OUTPUTS = {
    "evidence.jsonl",
    "results.json",
    "detail.xlsx",
    "summary.xlsx",
    "report.md",
}

PROFILE_CASES = [
    pytest.param(
        ["--profile", "windows", "--input", "tests/fixtures/windows/paste/good-collector.json"],
        64,
        id="windows",
    ),
    pytest.param(
        ["--profile", "linux", "--input", "tests/fixtures/linux_server/good.json"],
        8,
        id="linux",
    ),
    pytest.param(
        ["--profile", "unix", "--unix", "aix", "--input", "tests/fixtures/unix_server/aix/good.json"],
        67,
        id="unix-aix",
    ),
    pytest.param(
        ["--profile", "dbms", "--dbms", "postgresql", "--input", "tests/fixtures/dbms/postgresql/good.json"],
        26,
        id="dbms-postgresql",
    ),
    pytest.param(
        ["--profile", "dbms", "--dbms", "mysql", "--input", "tests/fixtures/dbms/mysql/good.json"],
        26,
        id="dbms-mysql",
    ),
    pytest.param(
        ["--profile", "dbms", "--dbms", "mariadb", "--input", "tests/fixtures/dbms/mariadb/good.json"],
        26,
        id="dbms-mariadb",
    ),
    pytest.param(
        ["--profile", "network", "--vendor", "cisco_ios", "--input", "tests/fixtures/network/cisco_ios/good.txt"],
        38,
        id="network-cisco-ios",
    ),
    pytest.param(
        ["--profile", "network", "--vendor", "junos", "--input", "tests/fixtures/network/junos/display_set_good.txt"],
        38,
        id="network-junos",
    ),
    pytest.param(
        ["--profile", "security-appliance", "--appliance-type", "firewall", "--input", "tests/fixtures/security_appliance/vulnerable.txt"],
        23,
        id="security-appliance",
    ),
]


@pytest.mark.parametrize(("arguments", "expected_count"), PROFILE_CASES)
def test_supported_profiles_create_complete_output_bundle(tmp_path: Path, arguments: list[str], expected_count: int):
    output = tmp_path / "output"
    result = runner.invoke(app, ["classify-file", *arguments, "--output", str(output)])

    assert result.exit_code == 0, result.output
    assert {path.name for path in output.iterdir()} == SEVEN_OUTPUTS

    payload = json.loads((output / "results.json").read_text(encoding="utf-8"))
    assert len(payload["results"]) == expected_count
    assert {row["status"] for row in payload["results"]} <= {
        "GOOD",
        "VULNERABLE",
        "MANUAL_REQUIRED",
    }

    detail = load_workbook(output / "detail.xlsx", read_only=True)
    summary = load_workbook(output / "summary.xlsx", read_only=True)
    assert detail["04_상세결과"].max_row == expected_count + 1
    assert detail["03_분야별통계"].max_row > 1
    assert detail["08_AI판정로그"].max_row == expected_count + 1
    assert summary["종합통계"].max_row > 1
    assert summary["분야별통계"].max_row > 1
    assert summary["권고문통계"].max_row > 1

    if any(row["status"] == "VULNERABLE" for row in payload["results"]):
        assert detail["05_취약점목록"].max_row > 1
    if any(row["status"] == "MANUAL_REQUIRED" for row in payload["results"]):
        assert detail["06_수동확인"].max_row > 1

    _assert_no_damaged_or_sensitive_output(output)


def test_no_advisory_creates_five_outputs(tmp_path: Path):
    output = tmp_path / "no-advisory"
    result = runner.invoke(
        app,
        [
            "classify-file",
            "--profile",
            "linux",
            "--input",
            "tests/fixtures/linux_server/good.json",
            "--output",
            str(output),
            "--no-advisory",
        ],
    )

    assert result.exit_code == 0, result.output
    assert {path.name for path in output.iterdir()} == FIVE_OUTPUTS


def test_questionnaire_korean_and_legacy_headers_import(tmp_path: Path):
    korean = tmp_path / "questionnaire-korean.xlsx"
    legacy = tmp_path / "questionnaire-legacy.xlsx"

    export_result = runner.invoke(
        app,
        ["questionnaire", "export", "--profile", "security-appliance", "--output", str(korean)],
    )
    assert export_result.exit_code == 0, export_result.output

    workbook = load_workbook(korean)
    assert workbook.sheetnames == QUESTIONNAIRE_SHEETS
    assert len(workbook.sheetnames) == 11
    expected_headers = ["항목ID", "질문", "답변유형", "필요증적", "선택지", "판정규칙", "답변", "증적상태", "비고"]
    assert [cell.value for cell in workbook["02_계정관리"][1]] == expected_headers

    for sheet_name in QUESTIONNAIRE_SHEETS[2:9]:
        sheet = workbook[sheet_name]
        for row_number in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_number, column=1).value:
                sheet.cell(row=row_number, column=7).value = "예"
                sheet.cell(row=row_number, column=8).value = "provided"
    workbook.save(korean)

    korean_output = tmp_path / "korean-output"
    korean_result = runner.invoke(
        app,
        [
            "questionnaire",
            "import",
            "--profile",
            "security-appliance",
            "--input",
            str(korean),
            "--output",
            str(korean_output),
        ],
    )
    assert korean_result.exit_code == 0, korean_result.output
    assert len(json.loads((korean_output / "results.json").read_text(encoding="utf-8"))["results"]) == 23

    workbook = load_workbook(korean)
    legacy_headers = [
        "item_id",
        "question",
        "answer_type",
        "required_evidence",
        "options",
        "mapping_rule",
        "answer",
        "evidence_status",
        "notes",
    ]
    for sheet_name in QUESTIONNAIRE_SHEETS[2:9]:
        sheet = workbook[sheet_name]
        for column, header in enumerate(legacy_headers, start=1):
            sheet.cell(row=1, column=column).value = header
    workbook.save(legacy)

    legacy_output = tmp_path / "legacy-output"
    legacy_result = runner.invoke(
        app,
        [
            "questionnaire",
            "import",
            "--profile",
            "security-appliance",
            "--input",
            str(legacy),
            "--output",
            str(legacy_output),
        ],
    )
    assert legacy_result.exit_code == 0, legacy_result.output
    assert len(json.loads((legacy_output / "results.json").read_text(encoding="utf-8"))["results"]) == 23


def test_support_matrix_is_linked_and_runtime_resources_are_synchronized():
    readme = Path("README.md").read_text(encoding="utf-8")
    matrix = Path("docs/END_TO_END_SUPPORT_MATRIX.md").read_text(encoding="utf-8")

    assert "docs/END_TO_END_SUPPORT_MATRIX.md" in readme
    for expected in [
        "Windows Server",
        "Unix Server",
        "PostgreSQL, MySQL, MariaDB",
        "Cisco IOS",
        "Juniper Junos",
        "Security Appliance",
        "MANUAL_REQUIRED",
        "command-response simulator",
    ]:
        assert expected in matrix

    pairs = [
        ("rulepacks/kcii-2025-12/unix.yaml", "src/kcii_audit/resources/rulepacks/kcii-2025-12/unix.yaml"),
        ("rulepacks/kcii-2025-12/unix_items_manifest.yaml", "src/kcii_audit/resources/rulepacks/kcii-2025-12/unix_items_manifest.yaml"),
        ("rulepacks/kcii-2025-12/security_appliance.yaml", "src/kcii_audit/resources/rulepacks/kcii-2025-12/security_appliance.yaml"),
        ("rulepacks/kcii-2025-12/security_appliance_items_manifest.yaml", "src/kcii_audit/resources/rulepacks/kcii-2025-12/security_appliance_items_manifest.yaml"),
        ("questionnaires/security_appliance_schema.yaml", "src/kcii_audit/resources/questionnaires/security_appliance_schema.yaml"),
        ("questionnaires/templates/security_appliance_questionnaire.xlsx", "src/kcii_audit/resources/questionnaires/templates/security_appliance_questionnaire.xlsx"),
    ]
    for root_path, resource_path in pairs:
        assert _sha256(Path(root_path)) == _sha256(Path(resource_path))


def _assert_no_damaged_or_sensitive_output(output: Path) -> None:
    combined = ""
    for path in output.iterdir():
        if path.suffix in {".json", ".jsonl", ".md"}:
            combined += path.read_text(encoding="utf-8")
        if path.suffix == ".xlsx":
            workbook = load_workbook(path, read_only=True)
            combined += "\n".join(
                str(cell.value or "")
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
            )

    assert "???" not in combined
    for forbidden in [
        "fw-prod-01.example.invalid",
        "198.51.100.80",
        "APPLIANCE_ADMIN",
        "not-a-real-token",
    ]:
        assert forbidden not in combined


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()
