from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from kcii_audit.cli import app
from kcii_audit.reports.questionnaire_writer import QUESTIONNAIRE_SHEETS

runner = CliRunner()


def _fill_answers(path: Path, answer: str, evidence_status: str = "provided") -> None:
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        if sheet.title not in QUESTIONNAIRE_SHEETS[2:9]:
            continue
        headers = [cell.value for cell in sheet[1]]
        answer_col = headers.index("answer") + 1
        evidence_col = headers.index("evidence_status") + 1
        for row in range(2, sheet.max_row + 1):
            if not sheet.cell(row=row, column=1).value:
                continue
            sheet.cell(row=row, column=answer_col).value = answer
            sheet.cell(row=row, column=evidence_col).value = evidence_status
    workbook.save(path)


def test_security_appliance_questionnaire_template_exists_with_required_sheets():
    workbook = load_workbook("questionnaires/templates/security_appliance_questionnaire.xlsx")

    assert workbook.sheetnames == QUESTIONNAIRE_SHEETS


def test_security_appliance_questionnaire_export_cli_creates_template(tmp_path):
    output = tmp_path / "security_appliance_questionnaire.xlsx"
    result = runner.invoke(
        app,
        [
            "questionnaire",
            "export",
            "--profile",
            "security-appliance",
            "--output",
            str(output),
        ],
    )

    assert result.exit_code == 0, result.output
    workbook = load_workbook(output)
    assert workbook.sheetnames == QUESTIONNAIRE_SHEETS
    assert workbook["02_계정관리"].max_row > 1


def test_security_appliance_questionnaire_import_blank_answers_are_manual_required(tmp_path):
    questionnaire = tmp_path / "security_appliance_questionnaire.xlsx"
    output_dir = tmp_path / "security-appliance"
    runner.invoke(
        app,
        ["questionnaire", "export", "--profile", "security-appliance", "--output", str(questionnaire)],
    )

    result = runner.invoke(
        app,
        [
            "questionnaire",
            "import",
            "--profile",
            "security-appliance",
            "--input",
            str(questionnaire),
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads((output_dir / "results.json").read_text(encoding="utf-8"))
    assert len(payload["results"]) == 23
    assert {row["status"] for row in payload["results"]} == {"MANUAL_REQUIRED"}
    advisory_text = (output_dir / "security_advisory.md").read_text(encoding="utf-8")
    assert "S-01" in advisory_text


def test_security_appliance_questionnaire_import_vulnerable_answers_create_advisory(tmp_path):
    questionnaire = tmp_path / "security_appliance_questionnaire.xlsx"
    output_dir = tmp_path / "security-appliance-vulnerable"
    runner.invoke(
        app,
        ["questionnaire", "export", "--profile", "security-appliance", "--output", str(questionnaire)],
    )
    _fill_answers(questionnaire, "아니오")

    result = runner.invoke(
        app,
        [
            "questionnaire",
            "import",
            "--profile",
            "security-appliance",
            "--input",
            str(questionnaire),
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads((output_dir / "results.json").read_text(encoding="utf-8"))
    assert {row["status"] for row in payload["results"]} == {"VULNERABLE"}
    assert (output_dir / "security_advisory.md").exists()
    assert (output_dir / "security_advisory.xlsx").exists()


def test_security_appliance_classify_file_accepts_questionnaire_xlsx(tmp_path):
    questionnaire = tmp_path / "security_appliance_questionnaire.xlsx"
    output_dir = tmp_path / "security-appliance-classify-xlsx"
    runner.invoke(
        app,
        ["questionnaire", "export", "--profile", "security-appliance", "--output", str(questionnaire)],
    )
    _fill_answers(questionnaire, "예")

    result = runner.invoke(
        app,
        [
            "classify-file",
            "--profile",
            "security-appliance",
            "--input",
            str(questionnaire),
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads((output_dir / "results.json").read_text(encoding="utf-8"))
    assert {row["status"] for row in payload["results"]} == {"GOOD"}


def test_security_appliance_questionnaire_evidence_missing_is_warning(tmp_path):
    questionnaire = tmp_path / "security_appliance_questionnaire.xlsx"
    output_dir = tmp_path / "security-appliance-warning"
    runner.invoke(
        app,
        ["questionnaire", "export", "--profile", "security-appliance", "--output", str(questionnaire)],
    )
    _fill_answers(questionnaire, "예", evidence_status="missing")

    result = runner.invoke(
        app,
        [
            "questionnaire",
            "import",
            "--profile",
            "security-appliance",
            "--input",
            str(questionnaire),
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    evidence_text = (output_dir / "evidence.jsonl").read_text(encoding="utf-8")
    assert "evidence_missing" in evidence_text


def test_security_appliance_questionnaire_outputs_do_not_store_sensitive_notes(tmp_path):
    questionnaire = tmp_path / "security_appliance_questionnaire.xlsx"
    output_dir = tmp_path / "security-appliance-sensitive"
    runner.invoke(
        app,
        ["questionnaire", "export", "--profile", "security-appliance", "--output", str(questionnaire)],
    )
    _fill_answers(questionnaire, "아니오")
    workbook = load_workbook(questionnaire)
    sheet = workbook["02_계정관리"]
    headers = [cell.value for cell in sheet[1]]
    notes_col = headers.index("notes") + 1
    sheet.cell(row=2, column=notes_col).value = "fw-prod-01.example.invalid 198.51.100.80 [APPLIANCE_ADMIN]"
    workbook.save(questionnaire)

    result = runner.invoke(
        app,
        [
            "questionnaire",
            "import",
            "--profile",
            "security-appliance",
            "--input",
            str(questionnaire),
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    combined = (
        (output_dir / "evidence.jsonl").read_text(encoding="utf-8")
        + (output_dir / "results.json").read_text(encoding="utf-8")
        + (output_dir / "report.md").read_text(encoding="utf-8")
        + (output_dir / "security_advisory.md").read_text(encoding="utf-8")
    )
    assert "fw-prod-01.example.invalid" not in combined
    assert "198.51.100.80" not in combined
    assert "[APPLIANCE_ADMIN]" not in combined


def test_security_appliance_interview_sheet_adds_warning_without_storing_free_text(tmp_path):
    questionnaire = tmp_path / "security_appliance_questionnaire.xlsx"
    output_dir = tmp_path / "security-appliance-interview-warning"
    runner.invoke(
        app,
        ["questionnaire", "export", "--profile", "security-appliance", "--output", str(questionnaire)],
    )
    _fill_answers(questionnaire, "예")
    workbook = load_workbook(questionnaire)
    sheet = workbook["09_인터뷰결과"]
    headers = [cell.value for cell in sheet[1]]
    answer_col = headers.index("answer_summary") + 1
    warning_col = headers.index("validation_warning") + 1
    sheet.cell(row=2, column=answer_col).value = "fw-prod-01.example.invalid 198.51.100.80 [APPLIANCE_ADMIN]"
    sheet.cell(row=2, column=warning_col).value = "담당자 답변과 증적 불일치"
    workbook.save(questionnaire)

    result = runner.invoke(
        app,
        [
            "questionnaire",
            "import",
            "--profile",
            "security-appliance",
            "--input",
            str(questionnaire),
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    evidence_text = (output_dir / "evidence.jsonl").read_text(encoding="utf-8")
    assert "validation_warning" in evidence_text
    assert "fw-prod-01.example.invalid" not in evidence_text
    assert "198.51.100.80" not in evidence_text
    assert "[APPLIANCE_ADMIN]" not in evidence_text
