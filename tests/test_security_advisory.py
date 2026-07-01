from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from kcii_audit.cli import app
from kcii_audit.engine.evaluator import evaluate_evidence
from kcii_audit.parsers.linux_server import records_from_linux_paste
from kcii_audit.reports.advisory_writer import generate_security_advisories

runner = CliRunner()


def test_vulnerable_results_generate_security_advisories():
    text = Path("tests/fixtures/linux_server/vulnerable.json").read_text(encoding="utf-8")
    results = evaluate_evidence(records_from_linux_paste(text, asset_id="linux-sensitive-01"))

    advisories = generate_security_advisories(results)

    assert len(advisories) == 8
    assert advisories[0].advisory_id == "ADV-001"
    assert advisories[0].masked_asset == "[ASSET_001]"
    assert advisories[0].raw_evidence_hash.startswith("sha256:")
    assert all(advisory.masked for advisory in advisories)


def test_good_results_default_advisory_is_empty_summary(tmp_path):
    output_dir = tmp_path / "linux-good"
    result = runner.invoke(
        app,
        [
            "classify-file",
            "--profile",
            "linux",
            "--input",
            "tests/fixtures/linux_server/good.json",
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    advisory = (output_dir / "security_advisory.md").read_text(encoding="utf-8")
    assert "권고 대상 항목이 없습니다." in advisory
    assert (output_dir / "security_advisory.xlsx").exists()
    report = (output_dir / "report.md").read_text(encoding="utf-8")
    assert "보안 권고문 요약" in report


def test_include_good_advisory_option_creates_maintenance_advisories(tmp_path):
    output_dir = tmp_path / "linux-good"
    result = runner.invoke(
        app,
        [
            "classify-file",
            "--profile",
            "linux",
            "--input",
            "tests/fixtures/linux_server/good.json",
            "--output",
            str(output_dir),
            "--include-good-advisory",
        ],
    )

    assert result.exit_code == 0, result.output
    advisory = (output_dir / "security_advisory.md").read_text(encoding="utf-8")
    assert "유지관리 권고" in advisory
    assert "ADV-001" in advisory


def test_no_advisory_option_skips_advisory_files(tmp_path):
    output_dir = tmp_path / "linux-vulnerable"
    result = runner.invoke(
        app,
        [
            "classify-file",
            "--profile",
            "linux",
            "--input",
            "tests/fixtures/linux_server/vulnerable.json",
            "--output",
            str(output_dir),
            "--no-advisory",
        ],
    )

    assert result.exit_code == 0, result.output
    assert (output_dir / "evidence.jsonl").exists()
    assert (output_dir / "results.json").exists()
    assert (output_dir / "detail.xlsx").exists()
    assert (output_dir / "summary.xlsx").exists()
    assert (output_dir / "report.md").exists()
    assert not (output_dir / "security_advisory.md").exists()
    assert not (output_dir / "security_advisory.xlsx").exists()


def test_excel_outputs_include_security_advisory_sheets(tmp_path):
    output_dir = tmp_path / "linux-vulnerable"
    result = runner.invoke(
        app,
        [
            "classify-file",
            "--profile",
            "linux",
            "--input",
            "tests/fixtures/linux_server/vulnerable.json",
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    detail = load_workbook(output_dir / "detail.xlsx")
    summary = load_workbook(output_dir / "summary.xlsx")
    advisory = load_workbook(output_dir / "security_advisory.xlsx")

    assert "09_보안권고문" in detail.sheetnames
    assert "권고문통계" in summary.sheetnames
    assert "보안권고문" in advisory.sheetnames
    assert detail["09_보안권고문"].max_row > 1


def test_security_advisory_does_not_include_sensitive_fixture_values(tmp_path):
    output_dir = tmp_path / "linux-vulnerable"
    result = runner.invoke(
        app,
        [
            "classify-file",
            "--profile",
            "linux",
            "--input",
            "tests/fixtures/linux_server/vulnerable.json",
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    combined = (
        (output_dir / "security_advisory.md").read_text(encoding="utf-8")
        + (output_dir / "report.md").read_text(encoding="utf-8")
    )
    assert "adminroot" not in combined
    assert "linux-vulnerable-01" not in combined
    assert "192.0.2.20" not in combined
    assert "[ASSET_001]" in combined
