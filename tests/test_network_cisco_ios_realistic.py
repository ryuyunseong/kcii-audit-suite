from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from typer.testing import CliRunner

from kcii_audit.cli import app
from kcii_audit.engine.evaluator import evaluate_evidence
from kcii_audit.parsers.network_cisco_ios import records_from_cisco_ios_paste
from kcii_audit.schemas.result import AssessmentStatus

runner = CliRunner()

REALISTIC_FIXTURE = Path("tests/fixtures/network/cisco_ios/realistic/sanitized_lab_mixed.txt")
SENSITIVE_PLACEHOLDERS = [
    "[HOST_1]",
    "[IP_1]",
    "[USER_1]",
    "[COMMUNITY_1]",
    "[SERIAL_1]",
    "[DOMAIN_1]",
    "[PATH_1]",
    "[BANNER_1]",
    "[SECRET_1]",
    "[IMAGE_1]",
    "[PLATFORM_1]",
    "[WILDCARD_1]",
    "[DESC_1]",
    "[SIZE_1]",
    "[VIEW_1]",
]


def test_realistic_cisco_ios_fixture_keeps_full_manifest_and_expanded_auto_coverage():
    records = records_from_cisco_ios_paste(REALISTIC_FIXTURE.read_text(encoding="utf-8"), asset_id="network-realistic")
    results = evaluate_evidence(records)
    statuses = {result.item_id: result.status for result in results}
    counts = Counter(result.status for result in results)

    assert [record.item_id for record in records] == [f"N-{index:02d}" for index in range(1, 39)]
    assert counts[AssessmentStatus.GOOD] == 27
    assert counts[AssessmentStatus.MANUAL_REQUIRED] == 11
    assert counts[AssessmentStatus.VULNERABLE] == 0
    assert statuses["N-08"] == AssessmentStatus.GOOD
    assert statuses["N-19"] == AssessmentStatus.GOOD
    assert statuses["N-20"] == AssessmentStatus.GOOD
    assert statuses["N-27"] == AssessmentStatus.GOOD
    assert statuses["N-31"] == AssessmentStatus.GOOD
    assert statuses["N-34"] == AssessmentStatus.GOOD


def test_realistic_cisco_ios_classify_file_creates_outputs_without_sensitive_values(tmp_path):
    output_dir = tmp_path / "network-realistic"
    result = runner.invoke(
        app,
        [
            "classify-file",
            "--profile",
            "network",
            "--vendor",
            "cisco_ios",
            "--input",
            str(REALISTIC_FIXTURE),
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    for name in [
        "evidence.jsonl",
        "results.json",
        "detail.xlsx",
        "summary.xlsx",
        "report.md",
        "security_advisory.md",
        "security_advisory.xlsx",
    ]:
        assert (output_dir / name).exists()

    payload = json.loads((output_dir / "results.json").read_text(encoding="utf-8"))
    assert len(payload["results"]) == 38

    combined = "\n".join(_output_text_values(output_dir))
    for placeholder in SENSITIVE_PLACEHOLDERS:
        assert placeholder not in combined


def _output_text_values(output_dir: Path) -> Iterable[str]:
    for name in ["evidence.jsonl", "results.json", "report.md", "security_advisory.md"]:
        yield (output_dir / name).read_text(encoding="utf-8")

    for name in ["detail.xlsx", "summary.xlsx", "security_advisory.xlsx"]:
        workbook = load_workbook(output_dir / name, data_only=True)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is not None:
                        yield str(value)
