from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from typer.testing import CliRunner

from kcii_audit.cli import app
from kcii_audit.engine.evaluator import evaluate_evidence
from kcii_audit.parsers.network_junos import records_from_junos_paste
from kcii_audit.schemas.result import AssessmentStatus

runner = CliRunner()

FIXTURE_DIR = Path("tests/fixtures/network/junos")
SENSITIVE_PLACEHOLDERS = [
    "[HOST_1]",
    "[IP_1]",
    "[IP_2]",
    "[IP_3]",
    "[USER_1]",
    "[COMMUNITY_1]",
    "[BANNER_1]",
    "[SECRET_1]",
    "[GROUP_1]",
    "[GROUP_2]",
    "[FILTER_1]",
    "[FILTER_2]",
    "[PREFIX_LIST_1]",
    "[PREFIX_LIST_2]",
]


def _fixture(name: str) -> str:
    return (FIXTURE_DIR / name).read_text(encoding="utf-8")


def test_junos_display_set_fixture_emits_full_network_manifest():
    records = records_from_junos_paste(_fixture("display_set_good.txt"), asset_id="network-junos")
    results = evaluate_evidence(records)
    statuses = {result.item_id: result.status for result in results}
    counts = Counter(result.status for result in results)

    assert [record.item_id for record in records] == [f"N-{index:02d}" for index in range(1, 39)]
    assert all(record.os_name == "Juniper Junos" for record in records)
    assert counts[AssessmentStatus.VULNERABLE] == 0
    assert statuses["N-01"] == AssessmentStatus.GOOD
    assert statuses["N-03"] == AssessmentStatus.GOOD
    assert statuses["N-06"] == AssessmentStatus.GOOD
    assert statuses["N-07"] == AssessmentStatus.GOOD
    assert statuses["N-08"] == AssessmentStatus.GOOD
    assert statuses["N-10"] == AssessmentStatus.GOOD
    assert statuses["N-11"] == AssessmentStatus.GOOD
    assert statuses["N-15"] == AssessmentStatus.GOOD
    assert statuses["N-16"] == AssessmentStatus.GOOD
    assert statuses["N-18"] == AssessmentStatus.GOOD
    assert statuses["N-19"] == AssessmentStatus.GOOD
    assert statuses["N-20"] == AssessmentStatus.GOOD
    assert statuses["N-21"] == AssessmentStatus.MANUAL_REQUIRED
    assert statuses["N-27"] == AssessmentStatus.MANUAL_REQUIRED
    assert statuses["N-32"] == AssessmentStatus.GOOD
    assert statuses["N-34"] == AssessmentStatus.GOOD


def test_junos_display_set_vulnerable_indicators_are_classified():
    records = records_from_junos_paste(_fixture("display_set_vulnerable.txt"), asset_id="network-junos")
    statuses = {result.item_id: result.status for result in evaluate_evidence(records)}

    assert statuses["N-03"] == AssessmentStatus.VULNERABLE
    assert statuses["N-07"] == AssessmentStatus.VULNERABLE
    assert statuses["N-08"] == AssessmentStatus.VULNERABLE
    assert statuses["N-18"] == AssessmentStatus.VULNERABLE
    assert statuses["N-19"] == AssessmentStatus.VULNERABLE
    assert statuses["N-20"] == AssessmentStatus.VULNERABLE
    assert statuses["N-21"] == AssessmentStatus.VULNERABLE
    assert statuses["N-26"] == AssessmentStatus.VULNERABLE
    assert statuses["N-27"] == AssessmentStatus.VULNERABLE
    assert statuses["N-31"] == AssessmentStatus.VULNERABLE
    assert statuses["N-33"] == AssessmentStatus.VULNERABLE


def test_junos_brace_style_config_fails_closed_to_manual_required():
    records = records_from_junos_paste(_fixture("brace_config_unsupported.txt"), asset_id="network-junos")
    results = evaluate_evidence(records)

    assert len(results) == 38
    assert {result.status for result in results} == {AssessmentStatus.MANUAL_REQUIRED}
    assert {record.evidence["collection_status"] for record in records} == {"needs_display_set"}
    assert all("display set" in record.evidence["reason"] for record in records)


def test_junos_realistic_display_set_handles_prompts_inactive_and_apply_groups():
    records = records_from_junos_paste(
        _fixture("realistic/display_set_realistic_sanitized.txt"),
        asset_id="network-junos-realistic",
    )
    results = evaluate_evidence(records)
    counts = Counter(result.status for result in results)

    assert [record.item_id for record in records] == [f"N-{index:02d}" for index in range(1, 39)]
    assert counts[AssessmentStatus.GOOD] == 14
    assert counts[AssessmentStatus.MANUAL_REQUIRED] == 24
    assert counts[AssessmentStatus.VULNERABLE] == 0
    assert all(record.evidence.get("inactive_lines_ignored") is True for record in records)
    assert all(record.evidence.get("inheritance_required") is True for record in records)
    manual_reasons = [
        record.evidence.get("reason", "")
        for record in records
        if record.evidence.get("manual_required")
    ]
    assert manual_reasons
    assert all("inheritance expansion" in reason for reason in manual_reasons)


def test_junos_xml_and_json_configs_fail_closed_to_manual_required():
    samples = [
        "<configuration><system><services><ssh /></services></system></configuration>",
        '{"configuration": {"system": {"services": {"ssh": {}}}}}',
    ]

    for sample in samples:
        records = records_from_junos_paste(sample, asset_id="network-junos-unsupported")
        results = evaluate_evidence(records)

        assert len(results) == 38
        assert {result.status for result in results} == {AssessmentStatus.MANUAL_REQUIRED}
        assert {record.evidence["collection_status"] for record in records} == {"unsupported_format"}


def test_junos_inheritance_fixtures_are_synthetic_and_sanitized():
    fixture_names = [
        "inheritance/display_set_with_apply_groups.txt",
        "inheritance/display_inheritance_effective_good.txt",
        "inheritance/display_inheritance_conflict.txt",
        "inheritance/display_inheritance_incomplete.txt",
    ]
    forbidden_patterns = [
        r"-----BEGIN (RSA |DSA |EC |OPENSSH )?PRIVATE KEY-----",
        r"AKIA[0-9A-Z]{16}",
        r"ASIA[0-9A-Z]{16}",
        r"ghp_[A-Za-z0-9_]{20,}",
        r"github_pat_[A-Za-z0-9_]+",
        r"sk-[A-Za-z0-9]{20,}",
        r"\b(?:\d{1,3}\.){3}\d{1,3}\b",
        r"(?i)\b(serial|license)\s+[A-Za-z0-9_-]{4,}",
    ]

    for name in fixture_names:
        text = _fixture(name)
        assert "[HOST_1]" in text
        assert "[USER_1]" in text
        for pattern in forbidden_patterns:
            assert not re.search(pattern, text), name


def test_junos_display_set_apply_groups_without_inheritance_context_stays_manual_required():
    records = records_from_junos_paste(
        _fixture("inheritance/display_set_with_apply_groups.txt"),
        asset_id="network-junos-inheritance-required",
    )
    results = evaluate_evidence(records)

    assert len(results) == 38
    assert [record.item_id for record in records] == [f"N-{index:02d}" for index in range(1, 39)]
    counts = Counter(result.status for result in results)

    assert counts[AssessmentStatus.VULNERABLE] == 0
    assert counts[AssessmentStatus.MANUAL_REQUIRED] >= 24
    assert all(record.evidence.get("inheritance_required") is True for record in records)
    assert all(record.evidence.get("input_format") == "junos_display_set" for record in records)
    assert all(
        "inheritance expansion" in record.evidence.get("reason", "")
        for record in records
        if record.evidence.get("manual_required")
    )


def test_junos_inheritance_only_fixtures_fail_closed_until_parser_support_exists():
    fixture_names = [
        "inheritance/display_inheritance_effective_good.txt",
        "inheritance/display_inheritance_conflict.txt",
        "inheritance/display_inheritance_incomplete.txt",
    ]

    for name in fixture_names:
        records = records_from_junos_paste(_fixture(name), asset_id=f"network-junos-{name}")
        results = evaluate_evidence(records)

        assert len(results) == 38
        assert [record.item_id for record in records] == [f"N-{index:02d}" for index in range(1, 39)]
        assert {result.status for result in results} == {AssessmentStatus.MANUAL_REQUIRED}
        assert {record.evidence["collection_status"] for record in records} == {"unsupported_output"}
        assert all("display-set configuration lines were not found" in record.evidence["reason"] for record in records)


def test_junos_classify_file_creates_outputs_without_sensitive_values(tmp_path):
    output_dir = tmp_path / "network-junos"
    result = runner.invoke(
        app,
        [
            "classify-file",
            "--profile",
            "network",
            "--vendor",
            "junos",
            "--input",
            str(FIXTURE_DIR / "display_set_good.txt"),
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0, result.output
    for name in [
        "evidence.jsonl",
        "results.json",
        "detail.xlsx",
        "summary.xlsx",
        "report.md",
        "security_advisory.md",
        "security_advisory.xlsx",
    ]:
        assert (output_dir / name).exists()

    payload = json.loads((output_dir / "results.json").read_text(encoding="utf-8"))
    assert len(payload["results"]) == 38

    combined = "\n".join(_output_text_values(output_dir))
    for placeholder in SENSITIVE_PLACEHOLDERS:
        assert placeholder not in combined


def _output_text_values(output_dir: Path) -> Iterable[str]:
    for name in ["evidence.jsonl", "results.json", "report.md", "security_advisory.md"]:
        yield (output_dir / name).read_text(encoding="utf-8")

    for name in ["detail.xlsx", "summary.xlsx", "security_advisory.xlsx"]:
        workbook = load_workbook(output_dir / name, data_only=True)
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for value in row:
                        if value is not None:
                            yield str(value)
        finally:
            workbook.close()
