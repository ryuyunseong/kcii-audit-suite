from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from kcii_audit.parsers.security_appliance_common import records_from_security_appliance_facts
from kcii_audit.parsers.security_appliance_items import SECURITY_APPLIANCE_FACT_BY_ITEM
from kcii_audit.schemas.evidence import EvidenceRecord

SKIPPED_SHEETS = {"00_작성안내", "01_자산목록", "10_자동판정"}
TRUE_ANSWERS = {"yes", "y", "true", "1", "예", "적용", "양호", "있음", "사용", "설정", "수행", "완료"}
FALSE_ANSWERS = {"no", "n", "false", "0", "아니오", "미적용", "취약", "없음", "미사용", "미설정", "미수행", "미완료"}
MISSING_EVIDENCE = {"missing", "none", "no", "n", "없음", "누락", "미제출", "미첨부"}
CONTRADICTORY_EVIDENCE = {"contradictory", "conflict", "inconsistent", "모순", "불일치"}

HEADER_ALIASES = {
    "항목ID": "item_id",
    "질문": "question",
    "답변유형": "answer_type",
    "필요증적": "required_evidence",
    "선택지": "options",
    "판정규칙": "mapping_rule",
    "답변": "answer",
    "증적상태": "evidence_status",
    "비고": "notes",
    "인터뷰주제": "interview_topic",
    "답변요약": "answer_summary",
    "검증경고": "validation_warning",
}


def records_from_security_appliance_questionnaire(
    path: Path,
    *,
    appliance_type: str = "firewall",
    asset_id: str = "security-appliance-questionnaire",
    guide_version: str = "kcii-2025-12",
) -> list[EvidenceRecord]:
    facts, warnings = parse_questionnaire_facts(path)
    raw_hash = "sha256:" + hashlib.sha256(path.read_bytes()).hexdigest()
    return records_from_security_appliance_facts(
        facts,
        appliance_type=appliance_type,
        asset_id=asset_id,
        guide_version=guide_version,
        raw_evidence_hash=raw_hash,
        evidence_source="questionnaire",
        warnings_by_item=warnings,
    )


def parse_questionnaire_facts(path: Path) -> tuple[dict[str, bool], dict[str, list[str]]]:
    workbook = load_workbook(path, data_only=True)
    facts: dict[str, bool] = {}
    warnings: dict[str, list[str]] = {}

    for sheet in workbook.worksheets:
        if sheet.title in SKIPPED_SHEETS:
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = _normalized_header(rows[0])
        if "item_id" not in header:
            continue

        for row in rows[1:]:
            item_id = _cell(row, header, "item_id")
            if item_id not in SECURITY_APPLIANCE_FACT_BY_ITEM:
                continue
            if sheet.title == "09_인터뷰결과":
                _parse_interview_warning(row, header, item_id, warnings)
                continue

            mapping_rule = _cell(row, header, "mapping_rule") or SECURITY_APPLIANCE_FACT_BY_ITEM[item_id]
            if mapping_rule not in SECURITY_APPLIANCE_FACT_BY_ITEM.values():
                _append_warning(warnings, item_id, "validation_warning")
                mapping_rule = SECURITY_APPLIANCE_FACT_BY_ITEM[item_id]

            answer = _cell(row, header, "answer")
            normalized_answer = _normalize_answer(answer)
            if normalized_answer is None:
                _append_warning(warnings, item_id, "missing_answer")
            else:
                facts[mapping_rule] = normalized_answer

            evidence_status = _cell(row, header, "evidence_status")
            normalized_evidence_status = str(evidence_status or "").strip().lower()
            if normalized_evidence_status in MISSING_EVIDENCE:
                _append_warning(warnings, item_id, "evidence_missing")
            if normalized_evidence_status in CONTRADICTORY_EVIDENCE:
                _append_warning(warnings, item_id, "validation_warning")

    return facts, warnings


def _parse_interview_warning(
    row: tuple[Any, ...],
    header: dict[str, int],
    item_id: str,
    warnings: dict[str, list[str]],
) -> None:
    evidence_status = _cell(row, header, "evidence_status")
    normalized_evidence_status = str(evidence_status or "").strip().lower()
    if normalized_evidence_status in MISSING_EVIDENCE:
        _append_warning(warnings, item_id, "evidence_missing")
    if normalized_evidence_status in CONTRADICTORY_EVIDENCE:
        _append_warning(warnings, item_id, "validation_warning")
    if _cell(row, header, "validation_warning"):
        _append_warning(warnings, item_id, "validation_warning")


def _normalized_header(row: tuple[Any, ...]) -> dict[str, int]:
    header: dict[str, int] = {}
    for index, value in enumerate(row):
        if value is None:
            continue
        name = str(value).strip()
        header[HEADER_ALIASES.get(name, name)] = index
    return header


def _cell(row: tuple[Any, ...], header: dict[str, int], name: str) -> str:
    index = header.get(name)
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def _normalize_answer(value: str) -> bool | None:
    normalized = value.strip().lower()
    if not normalized:
        return None
    if normalized in TRUE_ANSWERS:
        return True
    if normalized in FALSE_ANSWERS:
        return False
    return None


def _append_warning(warnings: dict[str, list[str]], item_id: str, warning: str) -> None:
    warnings.setdefault(item_id, []).append(warning)
