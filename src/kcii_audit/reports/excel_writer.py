from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook

from kcii_audit.reports.advisory_writer import ADVISORY_HEADERS, advisory_to_row, generate_security_advisories
from kcii_audit.schemas.result import EvaluationResult

DETAIL_SHEETS = [
    "00_표지",
    "01_자산목록",
    "02_종합통계",
    "03_분야별통계",
    "04_상세결과",
    "05_취약점목록",
    "06_수동확인",
    "07_예외및대체통제",
    "08_AI판정로그",
    "09_보안권고문",
]

DETAIL_HEADERS = [
    "자산ID",
    "자산명",
    "IP/식별자",
    "분야",
    "항목ID",
    "항목명",
    "중요도",
    "판정",
    "판정근거",
    "증적요약",
    "취약사유",
    "조치방안",
    "담당자",
    "조치기한",
    "재점검결과",
    "AI사용여부",
    "마스킹여부",
    "원본증적해시",
]


def write_detail_workbook(
    path: Path,
    results: list[EvaluationResult],
    *,
    include_good_advisory: bool = False,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = DETAIL_SHEETS[0]

    for sheet_name in DETAIL_SHEETS[1:]:
        workbook.create_sheet(sheet_name)

    cover = workbook["00_표지"]
    cover.append(["문서", "K-CII 상세결과"])
    cover.append(["상태", "scaffold placeholder"])

    asset_sheet = workbook["01_자산목록"]
    asset_sheet.append(["자산ID", "분야"])
    for result in results:
        asset_sheet.append([result.asset_id, result.platform])

    stats_sheet = workbook["02_종합통계"]
    stats_sheet.append(["판정", "건수"])
    for status, count in Counter(result.status_label for result in results).items():
        stats_sheet.append([status, count])

    detail_sheet = workbook["04_상세결과"]
    detail_sheet.append(DETAIL_HEADERS)
    for result in results:
        detail_sheet.append(
            [
                result.asset_id,
                result.asset_id,
                result.asset_id,
                result.platform,
                result.item_id,
                result.item_title,
                "",
                result.status_label,
                result.reason,
                result.evidence_summary,
                "",
                "",
                "",
                "",
                "",
                "Y" if result.ai_used else "N",
                "Y" if result.masked else "N",
                result.raw_evidence_hash or "",
            ]
        )

    manual_sheet = workbook["06_수동확인"]
    manual_sheet.append(["자산ID", "항목ID", "사유"])
    for result in results:
        if result.status.value == "MANUAL_REQUIRED":
            manual_sheet.append([result.asset_id, result.item_id, result.reason])

    advisory_sheet = workbook["09_보안권고문"]
    advisory_sheet.append(ADVISORY_HEADERS)
    for advisory in generate_security_advisories(results, include_good=include_good_advisory):
        advisory_sheet.append(advisory_to_row(advisory))

    workbook.save(path)


def write_summary_workbook(
    path: Path,
    results: list[EvaluationResult],
    *,
    include_good_advisory: bool = False,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "종합통계"
    summary.append(["판정", "건수"])
    for status, count in Counter(result.status_label for result in results).items():
        summary.append([status, count])

    by_domain = workbook.create_sheet("분야별통계")
    by_domain.append(["분야", "판정", "건수"])
    for (platform, status), count in Counter((result.platform, result.status_label) for result in results).items():
        by_domain.append([platform, status, count])

    advisory_stats = workbook.create_sheet("권고문통계")
    advisory_stats.append(["판정", "권고문수"])
    advisories = generate_security_advisories(results, include_good=include_good_advisory)
    for status, count in Counter(advisory.status_label for advisory in advisories).items():
        advisory_stats.append([status, count])
    if not advisories:
        advisory_stats.append(["권고 없음", 0])

    workbook.save(path)
